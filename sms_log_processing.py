import configparser
import os
import re
import openpyxl
import progressbar
from time import sleep
from progressbar import AnimatedMarker, Bar, BouncingBar, Counter, ETA, \
    FileTransferSpeed, FormatLabel, Percentage, \
    ProgressBar, ReverseBar, RotatingMarker, \
    SimpleProgress, Timer

def convert_excel_to_dict(wb, sheet_name, progressbar=None):
    """
    :param wb: class:`openpyxl.workbook.Workbook`
    :param sheet_name: characters
    :return: list of dict
    """
    header = []
    kpi_sum = []
    try:
        sheet_ranges = wb[sheet_name]

        for index, row in enumerate(sheet_ranges.iter_rows(row_offset=0)):
            if progressbar is not None:
                progressbar.update(index)
            sum_element = {'INDEX': index}
            for i in range(len(row)):
                if index == 0:
                    # header
                    if row[i].value is not None:
                        header.append(row[i].value.strip().upper())
                else:
                    # add expected data to kpi sum dictionary
                    if i < len(header):
                        sum_element[header[i]] = row[i].value
            if index > 0:
                kpi_sum.append(sum_element)
    except KeyError as keyerror:
        print(str(keyerror))
    return kpi_sum


if __name__ == '__main__':
    cf = configparser.ConfigParser()
    cf.read('setting.ini')
    template_folder = (cf['DEFAULT']['template_folder'])
    sms_log_folder = (cf['DEFAULT']['sms_log_folder'])

    sms_template_files = [x for x in os.listdir(template_folder) if os.path.isfile(os.path.join(template_folder, x))]
    sms_templates = []
    for f in sms_template_files:
        sms_template_workbook = openpyxl.load_workbook(os.path.join(template_folder, f))
        sms_templates.extend(convert_excel_to_dict(sms_template_workbook, sms_template_workbook.active.title))

    for tem in sms_templates:
        # print(tem['TEMPLATE'])
        # search_group = re.search(r'<.*?>', tem['TEMPLATE'])
        search_group = re.findall(r'<.*?>', tem['TEMPLATE'])
        tem['TEMPLATE'] = re.sub(r'<.*?>', r'[a-zA-Z0-9\s.]+', tem['TEMPLATE'])
        # print(search_group)
        # print(tem['TEMPLATE'])

    # sms = 'Generali chuc Be DO AN. HUY ngay Quoc te thieu nhi that vui ve va y nghia.'
    # # matchObj = (re.match(r'Generali chuc [a-zA-Z0-9]+ ngay Quoc te thieu nhi that vui ve va y nghia.', sms))
    # matchObj = (re.match(r'Generali chuc [a-zA-Z0-9\s.]+ ngay Quoc te thieu nhi that vui ve va y nghia.', sms))
    # if matchObj:
    #     print(matchObj.group())
    # else:
    #     print("No match!!")
    widgets1 = ['Sms_reading', progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()]
    widgets4 = ['Sms_processing', progressbar.Bar('=', '[', ']'), ' ', progressbar.Percentage()]

    sms_log_files = [x for x in os.listdir(sms_log_folder) if os.path.isfile(os.path.join(sms_log_folder, x))]
    bar1 = progressbar.ProgressBar(widgets=widgets1)

    sms_log = []
    bar1.start()
    for f in sms_log_files:
        sms_log_workbook = openpyxl.load_workbook(os.path.join(sms_log_folder, f))
        active_worksheet = sms_log_workbook.active
        bar1.maxval = active_worksheet.max_row
        sms_log.extend(convert_excel_to_dict(sms_log_workbook, sms_log_workbook.active.title, bar1))
    bar1.finish()

    bar = progressbar.ProgressBar(maxval=len(sms_log), widgets=widgets4)

    bar.start()
    for sms in sms_log:
        bar.update(sms['INDEX'])
        for template in sms_templates:
            try:
                # matchObj = (re.match(r'Generali chuc [a-zA-Z0-9\s.]+ ngay Quoc te thieu nhi that vui ve va y nghia.', sms['MESSAGE']))
                matchObj = (re.match(template['TEMPLATE'], sms['MESSAGE']))
                if matchObj:
                    # print('%s matched %s' % (matchObj.group(), template['TEMPLATE'],))
                    dept_column = 'K%s' % (sms['INDEX']+1)
                    sms_log_workbook.active[dept_column] = template['DEPT']
                    break
            except TypeError as typeError:
                print('%s: %s' % (typeError, sms, ))
    sms_log_workbook.active['K1'] = 'DEPT'
    sms_log_workbook.save(os.path.join(sms_log_folder, f))
    bar.finish()
